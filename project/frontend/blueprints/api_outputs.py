"""api_outputs.py - Output routes for CSV/file serving and report management."""

from __future__ import annotations

import csv
import io
import json
import shutil
from datetime import datetime
from pathlib import Path

from flask import Blueprint, jsonify, send_file, send_from_directory

import project.frontend.server_context as ctx
from project.frontend.api_errors import (
    ERR_INTERNAL,
    ERR_NOT_FOUND,
)

bp = Blueprint("api_outputs", __name__, url_prefix="/api/outputs")
RUN_STATE_FILE = ".registration_runs_state.json"
RUN_ARCHIVE_DIR = "archive/registration_runs"


def _outputs_root() -> Path:
    return ctx._job_output_dir(ctx._query_job_id())


def _run_state_path(outputs_root: Path | None = None) -> Path:
    root = outputs_root or _outputs_root()
    return root / RUN_STATE_FILE


def _load_run_state(outputs_root: Path | None = None) -> dict[str, object]:
    path = _run_state_path(outputs_root)
    if not path.exists():
        return {"pinned": []}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {"pinned": []}
    pinned = data.get("pinned", [])
    if not isinstance(pinned, list):
        pinned = []
    pinned = [Path(str(name)).name for name in pinned if str(name).strip()]
    return {"pinned": pinned}


def _save_run_state(state: dict[str, object], outputs_root: Path | None = None) -> None:
    root = outputs_root or _outputs_root()
    root.mkdir(parents=True, exist_ok=True)
    pinned = state.get("pinned", [])
    if not isinstance(pinned, list):
        pinned = []
    data = {"pinned": [Path(str(name)).name for name in pinned if str(name).strip()]}
    _run_state_path(root).write_text(json.dumps(data, indent=2), encoding="utf-8")


def _pin_run(run_name: str, outputs_root: Path | None = None) -> list[str]:
    state = _load_run_state(outputs_root)
    pinned = [name for name in state.get("pinned", []) if name != run_name]
    pinned.insert(0, run_name)
    state["pinned"] = pinned
    _save_run_state(state, outputs_root)
    return pinned


def _unpin_run(run_name: str, outputs_root: Path | None = None) -> list[str]:
    state = _load_run_state(outputs_root)
    pinned = [name for name in state.get("pinned", []) if name != run_name]
    state["pinned"] = pinned
    _save_run_state(state, outputs_root)
    return pinned


def _load_metrics_csv(path: Path) -> dict[str, float]:
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader, None)
        out: dict[str, float] = {}
        for row in reader:
            if len(row) < 2:
                continue
            try:
                out[str(row[0])] = float(row[1])
            except Exception:
                continue
        return out


def _verdict(metrics: dict[str, float], pre_metrics: dict[str, float]) -> tuple[str, str, str]:
    if not pre_metrics:
        return (
            "Needs visual check",
            "Only final metrics are available. Check the final overview image directly.",
            "neutral",
        )

    ncc_up = metrics.get("NCC", 0.0) >= pre_metrics.get("NCC", 0.0)
    ssim_up = metrics.get("SSIM", 0.0) >= pre_metrics.get("SSIM", 0.0)
    dice_ok = metrics.get("Dice", 0.0) >= (pre_metrics.get("Dice", 0.0) - 0.005)
    psnr_up = metrics.get("PSNR", 0.0) >= pre_metrics.get("PSNR", 0.0)
    mse_down = metrics.get("MSE", 1e9) <= pre_metrics.get("MSE", 1e9)

    good = sum([ncc_up, ssim_up, dice_ok, psnr_up, mse_down])
    if good >= 4:
        return (
            "Looks improved",
            "Most quality metrics moved in the right direction. Check the final overview first.",
            "good",
        )
    if good >= 3:
        return (
            "Probably usable",
            "Metric changes are mostly positive, but you should still inspect the final overview.",
            "warn",
        )
    return (
        "Needs review",
        "Metric changes are mixed. Use the before/final comparison before accepting the run.",
        "bad",
    )


def _registration_run_dirs(outputs_root: Path | None = None) -> list[Path]:
    root = outputs_root or _outputs_root()
    if not root.exists():
        return []

    runs: list[Path] = []
    for child in root.iterdir():
        if not child.is_dir():
            continue
        if (child / "registration_metadata.json").exists() and (
            child / "registration_metrics.csv"
        ).exists():
            runs.append(child)
    state = _load_run_state(root)
    pinned_order = [name for name in state.get("pinned", []) if isinstance(name, str)]
    runs_by_name = {run.name: run for run in runs}
    pinned_runs = [runs_by_name[name] for name in pinned_order if name in runs_by_name]
    other_runs = [run for run in runs if run.name not in pinned_order]
    other_runs.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return pinned_runs + other_runs


def _safe_registration_run_dir(run_name: str) -> Path | None:
    root = _outputs_root()
    safe = Path(run_name).name
    run_dir = root / safe
    if safe != run_name:
        return None
    if not run_dir.exists() or not run_dir.is_dir():
        return None
    if not (run_dir / "registration_metadata.json").exists():
        return None
    return run_dir


def _run_file_url(run_name: str, filename: str) -> str:
    return f"/api/outputs/registration-run/{run_name}/{filename}"


def _existing_run_url(run_dir: Path, filename: str) -> str:
    path = run_dir / filename
    if path.exists() and path.is_file():
        return _run_file_url(run_dir.name, filename)
    return ""


def _serialize_registration_run(run_dir: Path) -> dict[str, object]:
    meta_path = run_dir / "registration_metadata.json"
    metrics_path = run_dir / "registration_metrics.csv"
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    metrics = _load_metrics_csv(metrics_path)
    pre_metrics = meta.get("metrics_before_laplacian", {}) or {}

    staining = meta.get("staining_stats", {}) or {}
    if not staining:
        staining_path = run_dir / "staining_stats.json"
        if staining_path.exists():
            try:
                staining = json.loads(staining_path.read_text(encoding="utf-8"))
            except Exception:
                staining = {}

    verdict_title, verdict_body, verdict_tone = _verdict(metrics, pre_metrics)
    updated_at = datetime.fromtimestamp(run_dir.stat().st_mtime).isoformat(timespec="seconds")
    input_source = str(meta.get("input_source", ""))
    input_name = Path(input_source).name if input_source else run_dir.name
    backend = str(meta.get("backend", "")).strip().upper()
    pipeline_label = backend or "UNKNOWN"
    if meta.get("laplacian_enabled"):
        pipeline_label = f"{pipeline_label} + Laplacian"
    state = _load_run_state(run_dir.parent)
    pinned = run_dir.name in state.get("pinned", [])

    return {
        "name": run_dir.name,
        "input_name": input_name,
        "input_source": input_source,
        "backend": str(meta.get("backend", "")),
        "pipeline_label": pipeline_label,
        "pinned": pinned,
        "laplacian_enabled": bool(meta.get("laplacian_enabled")),
        "hemisphere": str(meta.get("hemisphere", "")),
        "target_um": meta.get("target_um"),
        "updated_at": updated_at,
        "verdict_title": verdict_title,
        "verdict_body": verdict_body,
        "verdict_tone": verdict_tone,
        "metrics": metrics,
        "pre_metrics": pre_metrics,
        "staining_stats": staining,
        "artifacts": {
            "overview": _existing_run_url(run_dir, "overview.png"),
            "overview_before": _existing_run_url(run_dir, "overview_before.png"),
            "summary": _existing_run_url(run_dir, "registration_summary.txt"),
            "metadata": _existing_run_url(run_dir, "registration_metadata.json"),
            "metrics_csv": _existing_run_url(run_dir, "registration_metrics.csv"),
            "staining_json": _existing_run_url(run_dir, "staining_stats.json"),
            "report": _existing_run_url(run_dir, "report.html"),
        },
    }


def _archive_run_dir(run_dir: Path) -> Path:
    archive_root = run_dir.parent / RUN_ARCHIVE_DIR
    archive_root.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archived = archive_root / f"{stamp}_{run_dir.name}"
    counter = 1
    while archived.exists():
        archived = archive_root / f"{stamp}_{run_dir.name}_{counter}"
        counter += 1
    shutil.move(str(run_dir), str(archived))
    return archived


@bp.get("/leaf")
def outputs_leaf():
    fp = _outputs_root() / "cell_counts_leaf.csv"
    if not fp.exists():
        return jsonify({"ok": False, "error": "output not found", "error_code": ERR_NOT_FOUND}), 404
    return send_from_directory(fp.parent, fp.name)


@bp.get("/leaf/<channel>")
def outputs_leaf_channel(channel: str):
    fp = _outputs_root() / f"cell_counts_leaf_{channel}.csv"
    if not fp.exists():
        return ("", 204, {"Content-Type": "text/csv; charset=utf-8"})
    return send_from_directory(fp.parent, fp.name)


@bp.get("/hierarchy")
def outputs_hierarchy():
    fp = _outputs_root() / "cell_counts_hierarchy.csv"
    if not fp.exists():
        return jsonify(
            {"ok": False, "error": "hierarchy output not found", "error_code": ERR_NOT_FOUND}
        ), 404
    return send_from_directory(fp.parent, fp.name)


@bp.get("/registration-qc")
def outputs_registration_qc():
    fp = _outputs_root() / "slice_registration_qc.csv"
    if not fp.exists():
        return jsonify(
            {"ok": False, "error": "registration QC not found", "error_code": ERR_NOT_FOUND}
        ), 404
    return send_from_directory(fp.parent, fp.name)


@bp.get("/reg-slice-list")
def outputs_reg_slice_list():
    reg_dir = _outputs_root() / "registered_slices"
    if not reg_dir.exists():
        return jsonify({"ok": True, "files": [], "count": 0})
    files = sorted(reg_dir.glob("slice_*_overlay.png"))
    return jsonify({"ok": True, "files": [f.name for f in files], "count": len(files)})


@bp.get("/reg-slice/<filename>")
def outputs_reg_slice_file(filename: str):
    reg_dir = _outputs_root() / "registered_slices"
    safe = Path(filename).name
    fp = reg_dir / safe
    if not fp.exists() or not safe.endswith(".png"):
        return jsonify({"ok": False, "error": "file not found", "error_code": ERR_NOT_FOUND}), 404
    return send_from_directory(str(reg_dir), safe)


@bp.get("/file-list")
def outputs_file_list():
    outputs_root = _outputs_root()
    if not outputs_root.exists():
        return jsonify({"ok": True, "files": []})
    files = []
    for f in sorted(outputs_root.iterdir()):
        if f.is_file():
            files.append({"name": f.name, "size": f.stat().st_size, "ext": f.suffix.lower()})
    return jsonify({"ok": True, "files": files, "dir": str(outputs_root)})


@bp.get("/registration-runs")
def outputs_registration_runs():
    runs: list[dict[str, object]] = []
    for run_dir in _registration_run_dirs(_outputs_root()):
        try:
            runs.append(_serialize_registration_run(run_dir))
        except Exception:
            continue
    return jsonify({"ok": True, "runs": runs, "count": len(runs)})


@bp.post("/registration-run/<run_name>/pin")
def outputs_registration_run_pin(run_name: str):
    run_dir = _safe_registration_run_dir(run_name)
    if run_dir is None:
        return jsonify({"ok": False, "error": "run not found", "error_code": ERR_NOT_FOUND}), 404
    pinned = _pin_run(run_dir.name, run_dir.parent)
    return jsonify({"ok": True, "run": run_dir.name, "pinned": pinned})


@bp.post("/registration-run/<run_name>/delete-bad")
def outputs_registration_run_delete_bad(run_name: str):
    run_dir = _safe_registration_run_dir(run_name)
    if run_dir is None:
        return jsonify({"ok": False, "error": "run not found", "error_code": ERR_NOT_FOUND}), 404
    archived_dir = _archive_run_dir(run_dir)
    _unpin_run(run_dir.name, archived_dir.parent.parent)
    return jsonify(
        {
            "ok": True,
            "run": run_dir.name,
            "archived_to": str(archived_dir),
        }
    )


@bp.get("/registration-run/<run_name>/<filename>")
def outputs_registration_run_file(run_name: str, filename: str):
    run_dir = _safe_registration_run_dir(run_name)
    if run_dir is None:
        return jsonify({"ok": False, "error": "run not found", "error_code": ERR_NOT_FOUND}), 404

    safe = Path(filename).name
    fp = run_dir / safe
    if not fp.exists() or not fp.is_file():
        return jsonify({"ok": False, "error": "file not found", "error_code": ERR_NOT_FOUND}), 404
    return send_from_directory(str(run_dir), safe)


@bp.get("/named/<filename>")
def outputs_named(filename: str):
    safe = Path(filename).name
    outputs_root = _outputs_root()
    fp = outputs_root / safe
    if not fp.exists():
        return jsonify({"ok": False, "error": "file not found", "error_code": ERR_NOT_FOUND}), 404
    return send_from_directory(str(outputs_root), safe)


@bp.get("/z-continuity")
def outputs_z_continuity():
    """Return AP (best_z) time-series data for the Z-continuity chart in the QC tab.

    Merges slice_registration_qc.csv with z_smoothness_report.json (if present)
    to return per-slice original, smoothed, and outlier flags.
    """
    outputs_root = _outputs_root()
    qc_csv = outputs_root / "slice_registration_qc.csv"
    if not qc_csv.exists():
        return jsonify(
            {
                "ok": False,
                "error": "slice_registration_qc.csv not found",
                "error_code": ERR_NOT_FOUND,
            }
        ), 404

    try:
        import pandas as pd

        df = pd.read_csv(qc_csv, usecols=lambda c: c in {"slice_id", "best_z", "registration_ok"})
        df = df.sort_values("slice_id").reset_index(drop=True)
        slices = df["slice_id"].tolist()
        original_z = [int(v) for v in df["best_z"].tolist()]
        reg_ok = [bool(v) for v in df.get("registration_ok", [True] * len(slices)).tolist()]
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc), "error_code": ERR_INTERNAL}), 500

    # Try to enrich with smoothness analysis
    smoothness_json = outputs_root / "z_smoothness_report.json"
    smoothed_z = original_z
    deviation = [0.0] * len(slices)
    is_outlier = [False] * len(slices)
    outlier_count = 0
    max_deviation = 0.0
    mean_deviation = 0.0
    monotone = True
    threshold = 8

    if smoothness_json.exists():
        try:
            report = json.loads(smoothness_json.read_text(encoding="utf-8"))
            # Align by slice_ids from report
            r_ids = report.get("slice_ids", [])
            r_sz = report.get("smoothed_z", [])
            r_dev = report.get("deviation", [])
            r_out = report.get("is_outlier", [])
            id_to_idx = {int(sid): i for i, sid in enumerate(r_ids)}

            smoothed_z = []
            deviation = []
            is_outlier = []
            for sid in slices:
                idx = id_to_idx.get(int(sid))
                if idx is not None:
                    smoothed_z.append(int(r_sz[idx]) if idx < len(r_sz) else 0)
                    deviation.append(float(r_dev[idx]) if idx < len(r_dev) else 0.0)
                    is_outlier.append(bool(r_out[idx]) if idx < len(r_out) else False)
                else:
                    smoothed_z.append(original_z[slices.index(sid)])
                    deviation.append(0.0)
                    is_outlier.append(False)

            outlier_count = int(report.get("outlier_count", 0))
            max_deviation = float(report.get("max_deviation", 0.0))
            mean_deviation = float(report.get("mean_deviation", 0.0))
            monotone = bool(report.get("monotone", True))
            threshold = int(report.get("max_dev_threshold", 8))
        except Exception:
            pass  # fallback to original values

    return jsonify(
        {
            "ok": True,
            "slice_ids": slices,
            "original_z": original_z,
            "smoothed_z": smoothed_z,
            "deviation": deviation,
            "is_outlier": is_outlier,
            "registration_ok": reg_ok,
            "outlier_count": outlier_count,
            "max_deviation": max_deviation,
            "mean_deviation": mean_deviation,
            "monotone": monotone,
            "threshold": threshold,
        }
    )


@bp.get("/qc-list")
def outputs_qc_list():
    qc_dir = _outputs_root() / "qc_overlays"
    if not qc_dir.exists():
        return jsonify({"ok": True, "files": [], "count": 0})
    files = sorted(qc_dir.glob("overlay_*.png"))
    return jsonify({"ok": True, "files": [f.name for f in files], "count": len(files)})


@bp.get("/qc-file/<filename>")
def outputs_qc_file(filename: str):
    qc_dir = _outputs_root() / "qc_overlays"
    safe = Path(filename).name
    return send_from_directory(str(qc_dir), safe)


@bp.get("/ap-density")
def outputs_ap_density():
    """Return per-AP-slice cell count for the AP density chart in the Results tab.

    Merges cell_counts_leaf.csv (has slice_id) with slice_registration_qc.csv
    (has slice_id → best_z / atlas AP coordinate).
    """
    import pandas as pd  # local import — only needed for this route

    outputs_root = _outputs_root()
    leaf_path = outputs_root / "cell_counts_leaf.csv"
    qc_path = outputs_root / "slice_registration_qc.csv"

    if not leaf_path.exists() or not qc_path.exists():
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "Required CSV files not found. Run the pipeline first.",
                    "error_code": ERR_NOT_FOUND,
                }
            ),
            404,
        )
    try:
        leaf = pd.read_csv(leaf_path)
        qc = pd.read_csv(qc_path)
    except Exception as exc:
        return jsonify(
            {"ok": False, "error": f"Failed to read CSVs: {exc}", "error_code": ERR_INTERNAL}
        ), 500

    # Identify the AP column — may be named best_z, atlas_z, ap_index, etc.
    ap_col = next((c for c in ("best_z", "atlas_z", "ap_index", "ap") if c in qc.columns), None)
    if ap_col is None or "slice_id" not in qc.columns or "slice_id" not in leaf.columns:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "Required columns not found in QC or leaf CSV.",
                    "error_code": ERR_NOT_FOUND,
                }
            ),
            400,
        )

    # Aggregate leaf counts by slice_id
    per_slice = leaf.groupby("slice_id", as_index=False)["count"].sum()
    # Join with QC to get AP coordinate
    merged = per_slice.merge(
        qc[["slice_id", ap_col]].drop_duplicates("slice_id"), on="slice_id", how="left"
    )
    merged = merged.dropna(subset=[ap_col]).sort_values(ap_col)
    merged[ap_col] = merged[ap_col].astype(int)

    ap_slices = [
        {
            "ap_index": int(row[ap_col]),
            "slice_id": int(row["slice_id"]),
            "cell_count": int(row["count"]),
        }
        for _, row in merged.iterrows()
    ]
    return jsonify({"ok": True, "ap_slices": ap_slices, "total_slices": len(ap_slices)})


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------


@bp.get("/excel")
def outputs_excel():
    """GET /api/outputs/excel?job=...
    Returns an .xlsx workbook with three sheets:
      - Hierarchy  : cell_counts_hierarchy.csv
      - Leaf       : cell_counts_leaf.csv
      - RunParams  : run_params_*.json (latest)
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
    except ImportError:
        return (
            jsonify({"ok": False, "error": "openpyxl not installed.", "error_code": ERR_INTERNAL}),
            500,
        )

    outputs_dir = _outputs_root()
    wb = Workbook()

    # ---- Sheet 1: Hierarchy ----
    hier_path = outputs_dir / "cell_counts_hierarchy.csv"
    if hier_path.exists():
        try:
            df_hier = pd.read_csv(hier_path, encoding="utf-8-sig")
            ws = wb.active
            ws.title = "Hierarchy"
            ws.append(list(df_hier.columns))
            for row in df_hier.itertuples(index=False):
                ws.append(list(row))
        except Exception:
            wb.active.title = "Hierarchy"
    else:
        wb.active.title = "Hierarchy"

    # ---- Sheet 2: Leaf ----
    leaf_path = outputs_dir / "cell_counts_leaf.csv"
    if leaf_path.exists():
        try:
            df_leaf = pd.read_csv(leaf_path, encoding="utf-8-sig")
            ws2 = wb.create_sheet("Leaf")
            ws2.append(list(df_leaf.columns))
            for row in df_leaf.itertuples(index=False):
                ws2.append(list(row))
        except Exception:
            wb.create_sheet("Leaf")
    else:
        wb.create_sheet("Leaf")

    # ---- Sheet 3: RunParams ----
    ws3 = wb.create_sheet("RunParams")
    params_files = sorted(outputs_dir.glob("run_params_*.json"), reverse=True)
    if params_files:
        try:
            params = json.loads(params_files[0].read_text(encoding="utf-8"))
            ws3.append(["Key", "Value"])
            for k, v in params.items():
                ws3.append([str(k), str(v)])
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="brainfast_results.xlsx",
    )


@bp.get("/coexpression")
def outputs_coexpression():
    """GET /api/outputs/coexpression?job=...
    Merges cell_counts_leaf_red.csv + cell_counts_leaf_green.csv per region.
    Returns {"ok": True, "regions": [{"acronym": ..., "name": ...,
             "count_red": ..., "count_green": ...}, ...]}
    """
    try:
        import pandas as pd
    except ImportError:
        return (
            jsonify({"ok": False, "error": "pandas not installed.", "error_code": ERR_INTERNAL}),
            500,
        )

    outputs_dir = _outputs_root()
    red_path = outputs_dir / "cell_counts_leaf_red.csv"
    green_path = outputs_dir / "cell_counts_leaf_green.csv"

    if not red_path.exists() and not green_path.exists():
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "No per-channel leaf CSVs found.",
                    "error_code": ERR_NOT_FOUND,
                }
            ),
            404,
        )

    def _load(p: Path, channel: str) -> pd.DataFrame | None:
        if not p.exists():
            return None
        try:
            df = pd.read_csv(p, encoding="utf-8-sig")
            count_col = next((c for c in ("count", "cell_count", "Count") if c in df.columns), None)
            if count_col is None:
                return None
            df = df.rename(columns={count_col: f"count_{channel}"})
            return df
        except Exception:
            return None

    red_df = _load(red_path, "red")
    green_df = _load(green_path, "green")

    key_cols = [
        c
        for c in ("acronym", "name", "id")
        if (
            (red_df is not None and c in red_df.columns)
            or (green_df is not None and c in green_df.columns)
        )
    ]
    if not key_cols:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "Cannot find region key columns.",
                    "error_code": ERR_INTERNAL,
                }
            ),
            500,
        )
    merge_key = key_cols[0]

    if red_df is not None and green_df is not None:
        merged = red_df.merge(
            green_df[[merge_key, "count_green"]], on=merge_key, how="outer"
        ).fillna(0)
    elif red_df is not None:
        merged = red_df.copy()
        merged["count_green"] = 0
    else:
        merged = green_df.copy()
        merged["count_red"] = 0

    out_cols = ["acronym", "name", "count_red", "count_green"]
    for c in out_cols:
        if c not in merged.columns:
            merged[c] = "" if c in ("acronym", "name") else 0

    regions = merged[out_cols].to_dict(orient="records")
    return jsonify(
        {
            "ok": True,
            "regions": regions,
            "channel_red_available": red_path.exists(),
            "channel_green_available": green_path.exists(),
        }
    )
